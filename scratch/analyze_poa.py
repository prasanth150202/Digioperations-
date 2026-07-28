import openpyxl

excel_path = r"C:\Users\Lenovo\Downloads\Monthly_Media_Buyer_POA_Template.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=False)

print("=== SHEETS IN WORKBOOK ===")
print(wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n==========================================")
    print(f"SHEET: {sheetname} (Dimensions: {ws.dimensions})")
    print(f"==========================================")
    
    # Iterate over all rows with content
    for r_idx, row in enumerate(ws.iter_rows(), start=1):
        row_vals = []
        for cell in row:
            if cell.value is not None:
                cell_info = f"{cell.coordinate}: {repr(cell.value)}"
                if cell.comment:
                    cell_info += f" [Comment: {cell.comment.text}]"
                if cell.data_type == 'f':
                    cell_info += " [Formula]"
                row_vals.append(cell_info)
        if row_vals:
            print(f"Row {r_idx:3d} | " + " | ".join(row_vals))

print("\n=== STYLES AND COLORS SUMMARY ===")
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    merged_ranges = [str(m) for m in ws.merged_cells.ranges]
    print(f"\nSheet {sheetname} Merged Cells ({len(merged_ranges)}):")
    for m in merged_ranges[:20]:
        print("  -", m)
    if len(merged_ranges) > 20:
        print(f"  ... and {len(merged_ranges)-20} more merged ranges")
