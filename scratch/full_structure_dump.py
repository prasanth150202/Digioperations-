import openpyxl

excel_path = r"C:\Users\Lenovo\Downloads\Monthly_Media_Buyer_POA_Template.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=False)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n========================================================")
    print(f"SHEET: {sheetname}")
    print(f"========================================================")
    
    # 1. Print Data Validations
    if ws.data_validations and ws.data_validations.dataValidation:
        print("Data Validations:")
        for dv in ws.data_validations.dataValidation:
            print(f"  - Sqref: {dv.sqref} | Type: {dv.type} | Formula1: {dv.formula1}")

    # 2. Print Headers and Rows (up to row 15)
    for r in range(1, 15):
        row_cells = []
        for c in range(1, 30):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                val = str(cell.value).replace('\n', ' ')
                if len(val) > 40:
                    val = val[:37] + "..."
                row_cells.append(f"{cell.coordinate}:{val}")
        if row_cells:
            print(f"Row {r:2d}: " + " | ".join(row_cells))
