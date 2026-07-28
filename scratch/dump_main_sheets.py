import openpyxl

excel_path = r"C:\Users\Lenovo\Downloads\Monthly_Media_Buyer_POA_Template.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

main_sheets = ['Overview', 'Communication', 'Competitors', 'Website Changes', 'Creative Plan', 'Retention Plan']

for sheetname in main_sheets:
    ws = wb[sheetname]
    print(f"\n========================================================")
    print(f"MAIN SHEET: {sheetname}")
    print(f"========================================================")
    
    for r in range(1, 15):
        row_vals = []
        for col_idx in range(1, 25):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                row_vals.append(f"{col_letter}{r}: {val}")
        if row_vals:
            print(f"Row {r:2d} | " + " | ".join(row_vals))
